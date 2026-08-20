"""CSV and Excel export for the admin reports."""
from __future__ import annotations

import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="7C2D12")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def to_csv(rows: list[dict], columns: list[tuple[str, str]]) -> bytes:
    """columns is a list of (key, Header Label) pairs."""
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow([row.get(key, "") for key, _ in columns])
    # utf-8-sig so Excel opens Indian rupee text and names correctly.
    return buffer.getvalue().encode("utf-8-sig")


def to_xlsx(
    sheets: list[tuple[str, list[dict], list[tuple[str, str]]]],
    title: str | None = None,
) -> bytes:
    """sheets is a list of (sheet_name, rows, columns)."""
    wb = Workbook()
    wb.remove(wb.active)

    for name, rows, columns in sheets:
        ws = wb.create_sheet(title=name[:31] or "Sheet")
        ws.append([label for _, label in columns])
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            ws.append([_coerce(row.get(key)) for key, _ in columns])

        for index, (key, label) in enumerate(columns, start=1):
            width = max(
                len(str(label)),
                *(len(str(row.get(key, ""))) for row in rows[:200] or [{}]),
            )
            ws.column_dimensions[get_column_letter(index)].width = min(max(width + 4, 12), 42)
        ws.freeze_panes = "A2"

    if title:
        wb.properties.title = title
    wb.properties.creator = "Annapurna Kitchen Billing Software"
    wb.properties.created = datetime.now()

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _coerce(value):
    """Numeric strings become real numbers so Excel can sum them."""
    if value is None or isinstance(value, (int, float)):
        return value
    text = str(value)
    try:
        return float(text) if ("." in text and text.replace(".", "", 1).lstrip("-").isdigit()) else (
            int(text) if text.lstrip("-").isdigit() else text
        )
    except (TypeError, ValueError):  # pragma: no cover - defensive
        return text


# --- column maps per report ---------------------------------------------

DAILY_COLUMNS = [
    ("date", "Date"),
    ("bill_count", "Bills"),
    ("discount", "Discount"),
    ("cgst", "CGST"),
    ("sgst", "SGST"),
    ("total", "Total Sales"),
]

PAYMENT_COLUMNS = [("mode", "Payment Mode"), ("count", "Count"), ("amount", "Amount")]

ITEM_COLUMNS = [
    ("name", "Item"),
    ("quantity_sold", "Qty Sold"),
    ("revenue", "Revenue"),
]

STAFF_COLUMNS = [
    ("name", "User"),
    ("username", "Username"),
    ("role", "Role"),
    ("bill_count", "Bills"),
    ("total_sales", "Total Sales"),
    ("discount_given", "Discount Given"),
    ("average_bill", "Average Bill"),
]

EXPENSE_COLUMNS = [
    ("date", "Date"),
    ("description", "Description"),
    ("category", "Category"),
    ("amount", "Amount"),
    ("created_by_name", "Recorded By"),
]

PNL_COLUMNS = [("metric", "Metric"), ("value", "Value")]

BILL_COLUMNS = [
    ("bill_number", "Bill No"),
    ("paid_at", "Paid At"),
    ("order_type", "Type"),
    ("table_number", "Table"),
    ("created_by_name", "Cashier"),
    ("customer_name", "Customer"),
    ("subtotal", "Subtotal"),
    ("discount_applied", "Discount"),
    ("cgst", "CGST"),
    ("sgst", "SGST"),
    ("total", "Total"),
]


def pnl_rows(report: dict) -> list[dict]:
    return [
        {"metric": "Gross revenue (incl. GST)", "value": report["gross_revenue"]},
        {"metric": "GST collected", "value": report["gst_collected"]},
        {"metric": "Net revenue (excl. GST)", "value": report["net_revenue"]},
        {"metric": "Discount given", "value": report["discount_given"]},
        {"metric": "Expenses", "value": report["expenses"]},
        {"metric": "Estimated profit / loss", "value": report["estimated_profit"]},
        {"metric": "Margin %", "value": report["margin_percentage"]},
        {"metric": "Bills", "value": report["bill_count"]},
    ]
