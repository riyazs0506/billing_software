"""Daily, item-wise, staff-wise, expense and P&L reports, plus exports."""
from __future__ import annotations

from datetime import date

from tests.conftest import data_of


def _sell(api, menu_ids, table_ids, table, item, quantity, mode="cash"):
    order = data_of(
        api.post("/api/orders", json={"order_type": "dine_in", "table_id": table_ids[table]})
    )
    api.post(
        "/api/orders/" + str(order["id"]) + "/items",
        json={"menu_item_id": menu_ids[item], "quantity": quantity},
    )
    bill = data_of(api.post("/api/billing/generate", json={"order_id": order["id"]}))["bill"]
    api.post(
        "/api/billing/bills/" + str(bill["id"]) + "/complete",
        json={"payments": [{"mode": mode, "amount": bill["total"]}]},
    )
    return bill


def test_daily_sales_totals(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 10, "cash")       # 210.00
    _sell(cashier_api, menu_ids, table_ids, "2", "Butter Chicken", 1, "upi")  # 378.00

    report = data_of(admin_api.get("/api/reports/daily"))
    assert report["summary"]["bill_count"] == 2
    assert report["summary"]["net_sales"] == "588.00"
    assert report["summary"]["gross_subtotal"] == "560.00"
    assert report["summary"]["gst_total"] == "28.00"
    assert report["summary"]["average_bill"] == "294.00"

    modes = {row["mode"]: row["amount"] for row in report["by_payment_mode"]}
    assert modes["cash"] == "210.00"
    assert modes["upi"] == "378.00"


def test_daily_sales_excludes_unpaid_bills(admin_api, cashier_api, menu_ids, table_ids):
    order = data_of(
        cashier_api.post(
            "/api/orders", json={"order_type": "dine_in", "table_id": table_ids["1"]}
        )
    )
    cashier_api.post(
        "/api/orders/" + str(order["id"]) + "/items",
        json={"menu_item_id": menu_ids["Chapati"], "quantity": 5},
    )
    cashier_api.post("/api/billing/generate", json={"order_id": order["id"]})

    report = data_of(admin_api.get("/api/reports/daily"))
    assert report["summary"]["bill_count"] == 0
    assert report["summary"]["net_sales"] == "0.00"


def test_item_wise_sales_and_best_sellers(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 12)
    _sell(cashier_api, menu_ids, table_ids, "2", "Butter Chicken", 2)

    report = data_of(admin_api.get("/api/reports/item-wise"))
    rows = {row["name"]: row for row in report["items"]}
    assert rows["Chapati"]["quantity_sold"] == 12
    assert rows["Chapati"]["revenue"] == "240.00"
    assert rows["Butter Chicken"]["quantity_sold"] == 2
    assert rows["Butter Chicken"]["revenue"] == "720.00"
    assert report["best_sellers"][0]["name"] == "Chapati"  # by quantity
    assert report["total_quantity"] == 14


def test_staff_wise_report_attributes_sales_to_the_signed_in_user(
    admin_api, cashier_api, menu_ids, table_ids
):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 10)
    _sell(admin_api, menu_ids, table_ids, "2", "Chapati", 5)

    report = data_of(admin_api.get("/api/reports/staff-wise"))
    by_username = {row["username"]: row for row in report["staff"]}
    assert by_username["cashier"]["bill_count"] == 1
    assert by_username["cashier"]["total_sales"] == "210.00"
    assert by_username["cashier"]["role"] == "cashier"
    assert by_username["admin"]["total_sales"] == "105.00"
    assert report["total_sales"] == "315.00"


def test_expense_report(admin_api):
    admin_api.post(
        "/api/expenses",
        json={"description": "Vegetables", "category": "Vegetables", "amount": "1200.00"},
    )
    admin_api.post(
        "/api/expenses", json={"description": "Gas cylinder", "category": "Gas", "amount": "1800.00"}
    )

    report = data_of(admin_api.get("/api/reports/expenses"))
    assert report["count"] == 2
    assert report["total"] == "3000.00"
    assert {row["category"] for row in report["by_category"]} == {"Vegetables", "Gas"}


def test_profit_and_loss(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Butter Chicken", 5)  # 1890.00 incl GST
    admin_api.post("/api/expenses", json={"description": "Chicken supply", "amount": "900.00"})

    report = data_of(admin_api.get("/api/reports/profit-loss"))
    assert report["gross_revenue"] == "1890.00"
    assert report["gst_collected"] == "90.00"
    assert report["net_revenue"] == "1800.00"
    assert report["expenses"] == "900.00"
    assert report["estimated_profit"] == "900.00"
    assert report["is_profit"] is True


def test_a_loss_is_reported_as_such(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 2)  # 40 net
    admin_api.post("/api/expenses", json={"description": "Rent", "amount": "5000.00"})

    report = data_of(admin_api.get("/api/reports/profit-loss"))
    assert report["is_profit"] is False
    assert report["estimated_profit"].startswith("-")


def test_dashboard_summarises_the_day(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 10, "cash")
    _sell(cashier_api, menu_ids, table_ids, "2", "Chapati", 10, "upi")
    admin_api.post("/api/expenses", json={"description": "Gas", "amount": "500.00"})

    dashboard = data_of(admin_api.get("/api/reports/dashboard"))
    assert dashboard["todays_bills"] == 2
    assert dashboard["todays_sales"] == "420.00"
    assert dashboard["cash_sales"] == "210.00"
    assert dashboard["upi_sales"] == "210.00"
    assert dashboard["card_sales"] == "0.00"
    assert dashboard["todays_expenses"] == "500.00"
    assert dashboard["best_sellers"][0]["name"] == "Chapati"
    assert len(dashboard["sales_trend"]) == 7
    assert dashboard["total_tables"] == 4


def test_dashboard_counts_active_tables_and_pending_bills(
    admin_api, cashier_api, menu_ids, table_ids
):
    order = data_of(
        cashier_api.post(
            "/api/orders", json={"order_type": "dine_in", "table_id": table_ids["5"]}
        )
    )
    cashier_api.post(
        "/api/orders/" + str(order["id"]) + "/items",
        json={"menu_item_id": menu_ids["Chapati"], "quantity": 3},
    )
    cashier_api.post("/api/billing/generate", json={"order_id": order["id"]})

    dashboard = data_of(admin_api.get("/api/reports/dashboard"))
    assert dashboard["active_tables"] == 1
    assert dashboard["pending_bills"] == 1


def test_date_range_filtering(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 4)
    today = date.today().isoformat()

    included = data_of(
        admin_api.get("/api/reports/daily?start_date=" + today + "&end_date=" + today)
    )
    assert included["summary"]["bill_count"] == 1

    excluded = data_of(
        admin_api.get("/api/reports/daily?start_date=2020-01-01&end_date=2020-01-31")
    )
    assert excluded["summary"]["bill_count"] == 0


def test_summary_bundles_every_report(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 4)
    summary = data_of(admin_api.get("/api/reports/summary"))
    assert set(summary) == {"daily", "item_wise", "staff_wise", "expenses", "profit_loss"}


# --- export --------------------------------------------------------------

def test_csv_export(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 6)
    response = admin_api.get("/api/reports/export/item-wise?format=csv")
    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    assert "attachment" in response.headers["Content-Disposition"]
    body = response.data.decode("utf-8-sig")
    assert "Item,Qty Sold,Revenue" in body
    assert "Chapati" in body


def test_xlsx_export(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 6)
    response = admin_api.get("/api/reports/export/daily?format=xlsx")
    assert response.status_code == 200
    assert response.data[:2] == b"PK"  # xlsx is a zip container
    assert ".xlsx" in response.headers["Content-Disposition"]


def test_export_all_bundles_every_sheet(admin_api, cashier_api, menu_ids, table_ids):
    _sell(cashier_api, menu_ids, table_ids, "1", "Chapati", 6)
    response = admin_api.get("/api/reports/export/all?format=xlsx")
    assert response.status_code == 200

    import io

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(response.data))
    assert "Item-wise Sales" in workbook.sheetnames
    assert "Staff-wise Sales" in workbook.sheetnames
    assert "Profit and Loss" in workbook.sheetnames


def test_unknown_report_is_rejected(admin_api):
    assert admin_api.get("/api/reports/export/nonsense").status_code == 422


def test_quick_ranges_are_offered(admin_api):
    ranges = data_of(admin_api.get("/api/reports/quick-ranges"))
    assert {r["label"] for r in ranges} >= {"Today", "Yesterday", "Last 7 days"}
